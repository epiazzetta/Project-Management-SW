# -*- coding: utf-8 -*-
# -------------------------------------------
# Project Management GUI - Version 2.3
# Author: Ermelino Piazzetta (modified)
# Using Tkinter with tabs and detailed views
# -------------------------------------------


import os
import platform
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import BarChart, Reference
import smtplib
from email.message import EmailMessage
from dotenv import load_dotenv

load_dotenv()

def list_existing_projects():
    # Retorna uma lista de nomes de projetos existentes (sem o prefixo 'project_' e a extensão '.xlsx')
    return [f[8:-5] for f in os.listdir() if f.startswith("project_") and f.endswith(".xlsx")]

def save_project_spreadsheet(project_name, items, totals_by_category):
    # Salva os itens do projeto e os totais por categoria em uma planilha Excel
    filename = f"project_{project_name}.xlsx"
    
    try:
        # Carrega a pasta de trabalho existente ou cria uma nova
        if os.path.exists(filename):
            wb = load_workbook(filename)
            # Se a aba "Project Items" já existe, a remove para criar uma nova limpa
            if "Project Items" in wb.sheetnames:
                wb.remove(wb["Project Items"])
            ws = wb.create_sheet(title="Project Items", index=1) # Cria como segunda aba (depois de Information)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Project Items" # Define o título da aba se for uma nova pasta de trabalho

        # Cabeçalho da tabela de itens
        ws.append(["Description", "Qty", "Unit", "Unit Price (R$)", "Total Price (R$)"])
        
        # Adiciona os itens à planilha
        for item in items:
            ws.append([item['desc'], item['qty'], item['unit'], item['unit_price'], item['total']])

        # Adiciona os totais por categoria
        # Certifica-se de que há espaço entre os itens e os totais
        if items:
            ws.append([]) # Linha em branco
        
        ws.append(["Totals by Category"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True) # Negrito
        ws.append(["Category", "Total"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True)


        # Guarda a linha inicial dos totais para o gráfico
        totals_header_row = ws.max_row
        for cat, tot in totals_by_category.items():
            ws.append([cat, tot])

        # Formatação
        bold = Font(bold=True)
        fill = PatternFill("solid", fgColor="BDD7EE")
        border = Border(*(Side(style='thin'),)*4)
        money = NamedStyle(name="money", number_format='"R$"#,##0.00')
        if "money" not in wb.named_styles:
            wb.add_named_style(money)
        
        # Aplica formatação a todas as células
        for r_idx, r in enumerate(ws.iter_rows(min_row=1, max_col=5, max_row=ws.max_row)):
            for cell in r:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                if r_idx == 0: # Cabeçalho da tabela de itens
                    cell.font=bold
                    cell.fill=fill
                # Aplica formatação de dinheiro nas colunas 4 e 5 (Preço Unitário e Preço Total)
                if cell.column in (4,5):
                    cell.style="money"
                # Formata os totais (se a célula está na coluna 2 e a linha é posterior ao cabeçalho dos totais)
                if cell.column == 2 and cell.row > totals_header_row:
                     cell.style="money"
        
        # Ajusta a largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Gráfico (agora com base nos totais por categoria corretos)
        chart = BarChart()
        chart.title = "Totais por Categoria"
        chart.x_axis.title = "Categoria"
        chart.y_axis.title = "Total (R$)"

        # Referências para os dados do gráfico (a partir da linha de "Category", pulando "Total by Category")
        data_ref = Reference(ws, min_col=2, min_row=totals_header_row + 2, max_row=ws.max_row)
        cats_ref = Reference(ws, min_col=1, min_row=totals_header_row + 2, max_row=ws.max_row)

        chart.add_data(data_ref, titles_from_data=False) # No titles_from_data, add_data expects numbers
        chart.set_categories(cats_ref)
        
        # Remove gráficos antigos antes de adicionar o novo
        ws._charts.clear()
        ws.add_chart(chart, f"A{ws.max_row + 3}") # Adiciona o gráfico abaixo da tabela

        # Salva o arquivo
        wb.save(filename)
        
        return sum(totals_by_category.values())
    
    except PermissionError:
        raise PermissionError(f"O arquivo '{filename}' está em uso ou você não tem permissão para escrevê-lo. Por favor, feche-o e tente novamente.")
    except Exception as e:
        raise Exception(f"Erro inesperado ao salvar a planilha '{filename}': {e}")

def save_project_info_sheet(project_name, info):
    # Salva as informações gerais do projeto em uma aba separada
    filename = f"project_{project_name}.xlsx"
    try:
        wb = load_workbook(filename) if os.path.exists(filename) else Workbook()
        
        # Remove a aba "Information" se ela já existe para criar uma nova
        if "Information" in wb.sheetnames:
            wb.remove(wb["Information"])
        
        ws = wb.create_sheet(title="Information", index=0) # Cria como a primeira aba
        ws.append(["Field", "Value"])

        # Adiciona as informações gerais do projeto
        info_fields = {
            "project_name": "Nome do Projeto",
            "manager": "Gerente",
            "manager_email": "Email do Gerente",
            "start_date": "Data de Início",
            "end_date": "Data de Término",
            "est_cost": "Custo Estimado (R$)"
        }

        # Assegura que o nome do projeto está na info para ser salvo
        if "project_name" not in info:
            info["project_name"] = project_name
            
        for key, display_name in info_fields.items():
            value = info.get(key, "")
            if key == "est_cost" and isinstance(value, (int, float)):
                ws.append([display_name, f"R$ {value:.2f}"])
            else:
                ws.append([display_name, value])
        
        ws.append([]) # Linha em branco para separar
        ws.append(["Participant Name", "Email"])
        
        # Adiciona os participantes
        for p in info.get("participants", []):
            ws.append([p["name"], p["email"]])
        
        # Ajusta a largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(filename)
    except PermissionError:
        raise PermissionError(f"O arquivo '{filename}' está em uso ou você não tem permissão para escrevê-lo. Por favor, feche-o e tente novamente.")
    except Exception as e:
        raise Exception(f"Erro inesperado ao salvar a planilha de informações '{filename}': {e}")


def update_summary(project_name, total_cost):
    # Atualiza o arquivo de resumo geral dos projetos
    fn = "project_summary.xlsx"
    try:
        if os.path.exists(fn):
            wb = load_workbook(fn)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws.append(["Project", "Total Cost"])
        
        # Procura e atualiza o projeto existente ou adiciona um novo
        found = False
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value == project_name:
                ws.cell(row=row_idx, column=2).value = total_cost
                found = True
                break
        if not found:
            ws.append([project_name, total_cost])
        
        # Aplica formatação de dinheiro na coluna de custo total
        money_style = NamedStyle(name="money_summary", number_format='"R$"#,##0.00')
        if "money_summary" not in wb.named_styles:
            wb.add_named_style(money_style)

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=2).style = "money_summary"

        # Ajusta a largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        wb.save(fn)
    except PermissionError:
        raise PermissionError(f"O arquivo de resumo '{fn}' está em uso ou você não tem permissão para escrevê-lo. Por favor, feche-o e tente novamente.")
    except Exception as e:
        raise Exception(f"Erro inesperado ao salvar o arquivo de resumo '{fn}': {e}")


def send_emails(project_name, info):
    # Envia e-mails para os participantes do projeto
    from_addr = info.get('manager_email')
    pwd = os.getenv("EMAIL_PASSWORD") # A senha do email deve estar em um arquivo .env

    if not from_addr or not pwd:
        messagebox.showwarning("Aviso de Email", "Email do gerente ou senha do email não configurados (verifique o arquivo .env).")
        return
    
    if not info.get('participants'):
        # messagebox.showinfo("Aviso de Email", "Nenhum participante para enviar e-mail.")
        return # Não é um erro, apenas não há ninguém para notificar

    all_emails_sent = True
    for p in info.get('participants', []):
        if not p.get('email'):
            messagebox.showwarning("Aviso de Email", f"Participante '{p.get('name')}' não tem e-mail válido para envio.")
            all_emails_sent = False
            continue

        msg = EmailMessage()
        msg["Subject"] = f"Novo Projeto: {project_name}"
        msg["From"] = from_addr
        msg["To"] = p.get('email')
        msg.set_content(
            f"Olá {p.get('name')},\n\nVocê foi adicionado ao projeto \"{project_name}\".\n"
            f"Gerente: {info.get('manager')}\nData de Início: {info.get('start_date')}\n"
            f"Data de Término Estimada: {info.get('end_date')}\nCusto Estimado: R$ {info.get('est_cost', 0):.2f}\n"
            f"\nAtenciosamente,\nSua Equipe de Gerenciamento de Projetos"
        )
        try:
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(from_addr, pwd)
                smtp.send_message(msg)
            # messagebox.showinfo("Email Enviado", f"Email enviado com sucesso para {p.get('email')}.")
        except Exception as e:
            messagebox.showwarning("Erro de Email", f"Não foi possível enviar e-mail para {p.get('email')}: {e}")
            all_emails_sent = False
    
    if all_emails_sent:
        messagebox.showinfo("Emails Enviados", "Todos os e-mails para os participantes foram enviados com sucesso (se houver).")
    else:
        messagebox.showwarning("Emails Enviados com Problemas", "Alguns e-mails não puderam ser enviados. Verifique os avisos anteriores.")


class ProjectManagerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Gerenciador de Projetos v2.2")
        # Removido self.geometry("800x600") para auto-ajuste

        self.projects = [] # Nomes dos projetos carregados
        self.project_info = None # Informações do projeto atualmente selecionado
        self.project_items = [] # Itens do projeto atualmente selecionado

        self.create_menu()
        self.create_widgets()
        self.load_projects()

    def create_menu(self):
        menubar = tk.Menu(self)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Novo Projeto...", command=self.new_project_dialog)
        file_menu.add_command(label="Excluir Projeto", command=self.delete_project)
        file_menu.add_separator()
        file_menu.add_command(label="Sair", command=self.quit)
        menubar.add_cascade(label="Arquivo", menu=file_menu)

        self.config(menu=menubar)

    def create_widgets(self):
        self.main_frame = ttk.Frame(self)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Painel Esquerdo: Lista de Projetos
        left_frame = ttk.Frame(self.main_frame, width=200) # Mantém uma largura sugerida, mas flexível
        left_frame.pack(side="left", fill="y", padx=(0, 10))

        ttk.Label(left_frame, text="Projetos Existentes").pack(pady=5)
        self.project_listbox = tk.Listbox(left_frame, exportselection=False)
        self.project_listbox.pack(fill="y", expand=True)
        self.project_listbox.bind("<<ListboxSelect>>", self.on_project_select)

        btn_frame_left = ttk.Frame(left_frame)
        btn_frame_left.pack(pady=5)
        ttk.Button(btn_frame_left, text="Novo Projeto", command=self.new_project_dialog).pack(side="top", pady=2) # Changed to side="top" for better stacking on small screens
        ttk.Button(btn_frame_left, text="Excluir Projeto", command=self.delete_project).pack(side="top", pady=2) # Changed to side="top"

        # Painel Direito: Abas para Detalhes
        right_frame = ttk.Frame(self.main_frame)
        right_frame.pack(side="left", fill="both", expand=True)

        self.tabs = ttk.Notebook(right_frame)
        self.tabs.pack(fill="both", expand=True)

        # Aba 1: Informações do Projeto
        self.tab_info = ttk.Frame(self.tabs)
        self.tabs.add(self.tab_info, text="Informações do Projeto")

        self.info_tree = ttk.Treeview(self.tab_info, columns=("field", "value"), show="headings")
        self.info_tree.heading("field", text="Campo")
        self.info_tree.heading("value", text="Valor")
        # Ajusta larguras das colunas para serem mais flexíveis
        self.info_tree.column("field", width=120, stretch=tk.NO)
        self.info_tree.column("value", stretch=tk.YES)
        self.info_tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Aba 2: Participantes
        self.tab_participants = ttk.Frame(self.tabs)
        self.tabs.add(self.tab_participants, text="Participantes")

        self.part_tree = ttk.Treeview(self.tab_participants, columns=("name", "email"), show="headings")
        self.part_tree.heading("name", text="Nome")
        self.part_tree.heading("email", text="Email")
        # Ajusta larguras das colunas para serem mais flexíveis
        self.part_tree.column("name", width=150, stretch=tk.NO)
        self.part_tree.column("email", stretch=tk.YES)
        self.part_tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Aba 3: Itens do Projeto
        self.tab_items = ttk.Frame(self.tabs)
        self.tabs.add(self.tab_items, text="Itens do Projeto")

        self.items_tree = ttk.Treeview(self.tab_items, columns=("desc", "qty", "unit", "unit_price", "total"), show="headings")
        self.items_tree.heading("desc", text="Descrição")
        self.items_tree.heading("qty", text="Quantidade")
        self.items_tree.heading("unit", text="Unidade")
        self.items_tree.heading("unit_price", text="Preço Unitário")
        self.items_tree.heading("total", text="Preço Total")
        # Ajusta larguras das colunas para serem mais flexíveis
        self.items_tree.column("desc", width=150, stretch=tk.YES)
        self.items_tree.column("qty", width=60, stretch=tk.NO)
        self.items_tree.column("unit", width=60, stretch=tk.NO)
        self.items_tree.column("unit_price", width=90, stretch=tk.NO)
        self.items_tree.column("total", width=90, stretch=tk.NO)
        self.items_tree.pack(fill="both", expand=True, padx=5, pady=5)

        btn_items = ttk.Frame(self.tab_items)
        btn_items.pack(pady=5)
        ttk.Button(btn_items, text="Adicionar Item", command=self.add_item_dialog).pack(side="left", padx=5)
        ttk.Button(btn_items, text="Remover Item Selecionado", command=self.remove_selected_item).pack(side="left", padx=5)
        ttk.Button(btn_items, text="Salvar Itens", command=self.save_items).pack(side="left", padx=5)

        # Botão para salvar todas as informações do projeto
        save_all_button = ttk.Button(right_frame, text="Salvar Todas as Alterações", command=self.save_all_project_data)
        save_all_button.pack(pady=10)


    def load_projects(self):
        # Carrega a lista de projetos existentes na listbox
        self.projects = list_existing_projects()
        self.project_listbox.delete(0, "end")
        if self.projects:
            for p in self.projects:
                self.project_listbox.insert("end", p)
            # Seleciona o primeiro projeto automaticamente se houver
            self.project_listbox.selection_set(0) 
            self.on_project_select(None) # Dispara o evento de seleção para carregar os detalhes
        else:
            self.project_listbox.insert("end", "<Nenhum projeto>")
            self.clear_detail_views() # Limpa as telas de detalhe se não houver projetos

    def clear_detail_views(self):
        # Limpa as treeviews das abas de detalhe
        for tree in (self.info_tree, self.part_tree, self.items_tree):
            for item in tree.get_children():
                tree.delete(item)
        self.project_info = None
        self.project_items = []

    def on_project_select(self, event):
        # Evento acionado ao selecionar um projeto na listbox
        selection = self.project_listbox.curselection()
        if not selection:
            self.clear_detail_views()
            return
        
        project_name = self.project_listbox.get(selection[0])
        if project_name == "<Nenhum projeto>":
            self.clear_detail_views()
            return
        
        self.load_project_details(project_name)

    def load_project_details(self, project_name):
        # Carrega os detalhes de um projeto específico a partir do seu arquivo Excel
        filename = f"project_{project_name.replace(' ', '_')}.xlsx" # Usar nome normalizado para o arquivo
        if not os.path.exists(filename):
            messagebox.showerror("Erro", "Arquivo do projeto não encontrado.")
            self.clear_detail_views() # Limpa os detalhes se o arquivo não existe
            return

        wb = load_workbook(filename)

        # Carrega a aba de Informações
        self.project_info = {"project_name": project_name} # Inicia com o nome do projeto (original com espaços)
        self.info_tree.delete(*self.info_tree.get_children())
        self.part_tree.delete(*self.part_tree.get_children()) # Limpa a treeview de participantes também
        
        if "Information" in wb.sheetnames:
            ws_info = wb["Information"]
            participants_start_row = -1
            
            # Define o mapeamento reverso para carregar as informações
            # Isso deve corresponder aos 'display_name' em save_project_info_sheet
            reverse_info_map = {
                "Nome do Projeto": "project_name",
                "Gerente": "manager",
                "Email do Gerente": "manager_email",
                "Data de Início": "start_date",
                "Data de Término": "end_date",
                "Custo Estimado (R$)": "est_cost"
            }

            # Lê as informações gerais primeiro e encontra a linha dos participantes
            for row_idx, row_values in enumerate(ws_info.iter_rows(min_row=2, values_only=True)):
                if row_values[0] == "Participant Name": # Encontrou o cabeçalho dos participantes
                    participants_start_row = row_idx + 2 # A linha dos dados do primeiro participante
                    break
                
                display_name = row_values[0]
                value = row_values[1]

                if display_name in reverse_info_map:
                    field_key = reverse_info_map[display_name]
                    # Limpa "R$" do custo estimado ao carregar
                    if field_key == "est_cost" and isinstance(value, str) and value.startswith("R$"):
                        try:
                            # Remove "R$", espaços e substitui vírgula por ponto para float
                            value = float(value.replace("R$", "").replace(",", ".").strip())
                        except ValueError:
                            value = 0.0 # Define como 0.0 se a conversão falhar
                    
                    self.project_info[field_key] = value
                    self.info_tree.insert("", "end", values=(display_name, row_values[1])) # Insere o valor original na treeview
            
            # --- NOVO BLOCO PARA CARREGAR PARTICIPANTES ---
            self.project_info["participants"] = [] # Inicializa a lista de participantes
            if participants_start_row != -1:
                for row_values in ws_info.iter_rows(min_row=participants_start_row, values_only=True):
                    if row_values[0] is None: # Parar se encontrar uma linha vazia após os participantes
                        break
                    # Garante que há pelo menos duas colunas para nome e email
                    if len(row_values) >= 2 and row_values[0] is not None:
                        name = row_values[0]
                        email = row_values[1]
                        self.project_info["participants"].append({"name": name, "email": email})
                        self.part_tree.insert("", "end", values=(name, email))
            # --- FIM DO NOVO BLOCO ---
        else:
            # Se não houver aba de informações, garante que as trees de info/participantes estão vazias
            self.info_tree.delete(*self.info_tree.get_children())
            self.part_tree.delete(*self.part_tree.get_children())
            self.project_info["participants"] = [] # Garante que a lista está vazia


        # Carrega a aba de Itens do Projeto
        self.project_items = []
        self.items_tree.delete(*self.items_tree.get_children())

        if "Project Items" in wb.sheetnames:
            ws_items = wb["Project Items"]
            for row in ws_items.iter_rows(min_row=2, max_col=5, values_only=True):
                # Para de ler se encontrar uma linha vazia ou o cabeçalho dos totais
                if all(x is None for x in row) or (isinstance(row[0], str) and row[0].startswith("Totals by")):
                    break
                if row[0] is not None: # Verifica se a descrição não é None
                    self.project_items.append({
                        "desc": row[0],
                        "qty": row[1],
                        "unit": row[2],
                        "unit_price": row[3],
                        "total": row[4],
                    })
                    self.items_tree.insert("", "end", values=row)

    def new_project_dialog(self):
        # Abre o diálogo para criar um novo projeto
        dlg = NewProjectDialog(self)
        self.wait_window(dlg) # Espera o diálogo fechar
        
        # Verifica se o diálogo retornou resultados (se o botão "Save Project" foi clicado)
        if dlg.result:
            info = dlg.result['info']
            items = dlg.result['items']
            project_name = info.get('project_name')

            # Normaliza o nome do projeto para o arquivo
            project_name_for_file = project_name.replace(" ", "_")

            # Verifica se o nome do projeto já existe antes de tentar salvar
            if f"project_{project_name_for_file}.xlsx" in [f.lower() for f in os.listdir() if f.startswith("project_") and f.endswith(".xlsx")]:
                if not messagebox.askyesno("Projeto Existente", f"Já existe um projeto com o nome '{project_name}'. Deseja sobrescrever?"):
                    return # Não sobrescreve, aborta a criação do projeto
            
            # Calcula os totais por categoria
            totals = defaultdict(float)
            for item in items:
                totals[item['desc']] += item['total'] # Aqui 'desc' é usado como categoria para os totais

            try:
                # Salva as informações e os itens nos arquivos Excel
                # A ordem importa: primeiro info, depois itens.
                save_project_info_sheet(project_name_for_file, info)
                total_cost = save_project_spreadsheet(project_name_for_file, items, totals)
                
                # Atualiza o resumo geral e envia e-mails
                update_summary(project_name, total_cost) # Aqui usa o nome original para o resumo
                send_emails(project_name, info)
                
                messagebox.showinfo("Sucesso", f"Projeto '{project_name}' criado e salvo com sucesso!")
                self.load_projects() # Recarrega a lista principal para mostrar o novo projeto
                # Seleciona o projeto recém-criado na lista para que seus detalhes apareçam
                if project_name in self.projects:
                    index = self.projects.index(project_name)
                    self.project_listbox.selection_clear(0, tk.END)
                    self.project_listbox.selection_set(index)
                    self.project_listbox.see(index) # Garante que está visível
                    self.on_project_select(None) # Carrega os detalhes do novo projeto
                else:
                    # Isso pode acontecer se houver um problema no load_projects
                    messagebox.showwarning("Aviso", "Projeto salvo, mas não apareceu automaticamente na lista. Tente recarregar.")

            except PermissionError as pe:
                messagebox.showerror("Erro de Permissão", str(pe))
            except Exception as e:
                messagebox.showerror("Erro ao Criar Projeto", f"Ocorreu um erro ao criar o projeto: {e}")

    def delete_project(self):
        # Deleta um projeto selecionado
        sel = self.project_listbox.curselection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um projeto para deletar.")
            return
        
        project_name = self.project_listbox.get(sel[0])
        if project_name == "<Nenhum projeto>":
            messagebox.showwarning("Aviso", "Nenhum projeto válido selecionado para deletar.")
            return

        if messagebox.askyesno("Excluir Projeto", f"Tem certeza que deseja EXCLUIR o projeto '{project_name}'? Esta ação é irreversível e removerá o arquivo Excel associado."):
            # Usa o nome normalizado para o arquivo
            fn = f"project_{project_name.replace(' ', '_')}.xlsx"
            try:
                if os.path.exists(fn):
                    os.remove(fn)
                    update_summary(project_name, 0) # Atualiza o resumo para remover/zerar o projeto
                    messagebox.showinfo("Sucesso", f"Projeto '{project_name}' deletado com sucesso.")
                else:
                    messagebox.showwarning("Aviso", f"O arquivo do projeto '{project_name}.xlsx' não foi encontrado, mas o projeto será removido da lista.")

            except PermissionError as pe:
                messagebox.showerror("Erro de Permissão", str(pe))
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível deletar o projeto: {e}")
            finally:
                self.load_projects() # Recarrega a lista de projetos
                self.clear_detail_views() # Limpa a exibição de detalhes

    def add_item_dialog(self):
        # Abre o diálogo para adicionar um item ao projeto ATUALMENTE SELECIONADO
        if not self.project_info:
            messagebox.showwarning("Aviso", "Por favor, selecione ou crie um projeto primeiro para adicionar itens.")
            return

        dlg = AddItemDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            item = dlg.result
            # Verifica se já existe um item com a mesma descrição no projeto atual
            if any(i['desc'].lower() == item['desc'].lower() for i in self.project_items):
                messagebox.showwarning("Duplicado", "Um item com esta descrição já existe no projeto atual. Considere editar a quantidade ou usar uma descrição diferente.")
                return
            self.project_items.append(item)
            self.items_tree.insert("", "end", values=(item['desc'], item['qty'], item['unit'], item['unit_price'], item['total']))

    def remove_selected_item(self):
        # Remove um item selecionado da treeview e da lista interna
        sel = self.items_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um item para remover.")
            return
        
        if not self.project_info: # Verifica se há um projeto carregado
            messagebox.showwarning("Aviso", "Nenhum projeto ativo para remover itens.")
            return

        idx = self.items_tree.index(sel[0])
        self.items_tree.delete(sel[0])
        del self.project_items[idx]
        messagebox.showinfo("Item Removido", "Item removido da lista. Lembre-se de clicar em 'Salvar Itens' ou 'Salvar Todas as Alterações' para persistir a mudança.")


    def save_items(self):
        # Salva os itens do projeto atualmente carregado
        if not self.project_info:
            messagebox.showwarning("Aviso", "Nenhum projeto selecionado para salvar os itens.")
            return
        
        project_name = self.project_info.get('project_name')
        if not project_name:
            messagebox.showwarning("Aviso", "Informações do projeto inválidas para salvar itens.")
            return
        
        totals = defaultdict(float)
        for item in self.project_items:
            totals[item['desc']] += item['total'] # Calcula os totais baseados na descrição
        
        try:
            # Usa o nome normalizado para o arquivo
            save_project_spreadsheet(project_name.replace(" ", "_"), self.project_items, totals)
            update_summary(project_name, sum(totals.values()))
            messagebox.showinfo("Salvo", "Itens do projeto salvos com sucesso.")
            self.load_project_details(project_name) # Recarrega os detalhes para atualizar o gráfico etc.
        except PermissionError as pe:
            messagebox.showerror("Erro de Permissão", str(pe))
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar itens do projeto: {e}")

    def save_all_project_data(self):
        # Salva todas as informações do projeto atualmente carregado (info, participantes e itens)
        if not self.project_info:
            messagebox.showwarning("Aviso", "Nenhum projeto selecionado para salvar todas as alterações.")
            return

        project_name = self.project_info.get('project_name')
        if not project_name:
            messagebox.showwarning("Aviso", "Informações do projeto inválidas para salvar tudo.")
            return
        
        try:
            # Recalcula totais por categoria antes de salvar a planilha de itens
            totals = defaultdict(float)
            for item in self.project_items:
                totals[item['desc']] += item['total']

            # Salva informações e participantes (usando o nome normalizado para o arquivo)
            save_project_info_sheet(project_name.replace(" ", "_"), self.project_info)
            # Salva itens e atualiza resumo (usando o nome normalizado para o arquivo)
            save_project_spreadsheet(project_name.replace(" ", "_"), self.project_items, totals)
            update_summary(project_name, sum(totals.values()))

            messagebox.showinfo("Sucesso", f"Todas as informações do projeto '{project_name}' foram salvas.")
            self.load_project_details(project_name) # Recarrega para refletir quaisquer mudanças visíveis

        except PermissionError as pe:
            messagebox.showerror("Erro de Permissão", str(pe))
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao salvar todas as informações: {e}")


class NewProjectDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Novo Projeto")
        self.result = None # Vai armazenar os dados do projeto se salvo

        # Ajuste o tamanho inicial para ser menor e centralizado
        self.initial_width = 500
        self.initial_height = 600 # Um pouco menor para o diálogo
        self.center_window(self.initial_width, self.initial_height, parent)
        self.transient(parent)
        self.grab_set() # Torna este diálogo modal (bloqueia a janela principal)

        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, padx=10, pady=10)

        # Informações Gerais do Projeto
        info_frame = ttk.LabelFrame(frm, text="Informações do Projeto")
        info_frame.pack(fill="x", padx=5, pady=5) # pack para preenchimento horizontal

        ttk.Label(info_frame, text="Nome do Projeto:").pack(anchor="w", padx=5, pady=(5,0))
        self.entry_name = ttk.Entry(info_frame)
        self.entry_name.pack(fill="x", padx=5)

        ttk.Label(info_frame, text="Nome do Gerente:").pack(anchor="w", padx=5, pady=(5,0))
        self.entry_manager = ttk.Entry(info_frame)
        self.entry_manager.pack(fill="x", padx=5)

        ttk.Label(info_frame, text="Email do Gerente:").pack(anchor="w", padx=5, pady=(5,0))
        self.entry_manager_email = ttk.Entry(info_frame)
        self.entry_manager_email.pack(fill="x", padx=5)

        ttk.Label(info_frame, text="Data de Início (YYYY-MM-DD):").pack(anchor="w", padx=5, pady=(5,0))
        self.entry_start = ttk.Entry(info_frame)
        self.entry_start.pack(fill="x", padx=5)

        ttk.Label(info_frame, text="Data de Término Estimada (YYYY-MM-DD):").pack(anchor="w", padx=5, pady=(5,0))
        self.entry_end = ttk.Entry(info_frame)
        self.entry_end.pack(fill="x", padx=5)

        ttk.Label(info_frame, text="Custo Estimado (R$):").pack(anchor="w", padx=5, pady=(5,0))
        self.entry_cost = ttk.Entry(info_frame)
        self.entry_cost.pack(fill="x", padx=5)

        # Participantes
        part_frame = ttk.LabelFrame(frm, text="Participantes")
        part_frame.pack(fill="both", expand=True, padx=5, pady=5)

        self.participants = [] # Lista para armazenar dicionários de participantes

        self.part_tree = ttk.Treeview(part_frame, columns=("name", "email"), show="headings", height=3) # Altura mais reduzida
        self.part_tree.heading("name", text="Nome")
        self.part_tree.heading("email", text="Email")
        self.part_tree.column("name", width=120, stretch=tk.NO)
        self.part_tree.column("email", stretch=tk.YES)
        self.part_tree.pack(fill="both", expand=True, padx=5, pady=5)

        btn_part = ttk.Frame(part_frame)
        btn_part.pack(fill="x", padx=5, pady=2)
        ttk.Button(btn_part, text="Adicionar Participante", command=self.add_participant_dialog).pack(side="left", padx=5)
        ttk.Button(btn_part, text="Remover Selecionado", command=self.remove_selected_participant).pack(side="left", padx=5)

        # Itens do Projeto
        items_frame = ttk.LabelFrame(frm, text="Itens do Projeto")
        items_frame.pack(fill="both", expand=True, padx=5, pady=5)

        self.items = [] # Lista para armazenar dicionários de itens

        self.items_tree = ttk.Treeview(items_frame, columns=("desc", "qty", "unit", "unit_price", "total"), show="headings", height=3) # Altura mais reduzida
        self.items_tree.heading("desc", text="Descrição")
        self.items_tree.heading("qty", text="Quantidade")
        self.items_tree.heading("unit", text="Unidade")
        self.items_tree.heading("unit_price", text="Preço Unitário")
        self.items_tree.heading("total", text="Preço Total")
        self.items_tree.column("desc", width=100, stretch=tk.YES)
        self.items_tree.column("qty", width=50, stretch=tk.NO)
        self.items_tree.column("unit", width=50, stretch=tk.NO)
        self.items_tree.column("unit_price", width=70, stretch=tk.NO)
        self.items_tree.column("total", width=70, stretch=tk.NO)
        self.items_tree.pack(fill="both", expand=True, padx=5, pady=5)

        btn_items = ttk.Frame(items_frame)
        btn_items.pack(fill="x", padx=5, pady=2)
        ttk.Button(btn_items, text="Adicionar Item", command=self.add_item_dialog).pack(side="left", padx=5)
        ttk.Button(btn_items, text="Remover Selecionado", command=self.remove_selected_item).pack(side="left", padx=5)

        # Botões de Ação Final
        btn_frame_bottom = ttk.Frame(frm)
        btn_frame_bottom.pack(pady=10)
        ttk.Button(btn_frame_bottom, text="Salvar Projeto", command=self.save_project).pack(side="left", padx=10)
        ttk.Button(btn_frame_bottom, text="Cancelar", command=self.destroy).pack(side="left", padx=10)

    def center_window(self, width, height, parent):
        # Obtém as dimensões da tela
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # Calcula a posição para centralizar
        x = (screen_width / 2) - (width / 2)
        y = (screen_height / 2) - (height / 2)
        self.geometry(f'{int(width)}x{int(height)}+{int(x)}+{int(y)}')


    def add_participant_dialog(self):
        # Abre o diálogo para adicionar um participante ao novo projeto (ainda em criação)
        dlg = AddParticipantDialog(self)
        self.wait_window(dlg)
        if dlg.result:
            p = dlg.result
            if any(x['email'].lower() == p['email'].lower() for x in self.participants):
                messagebox.showwarning("Duplicado", "Já existe um participante com este e-mail neste projeto.")
                return
            self.participants.append(p)
            self.part_tree.insert("", "end", values=(p["name"], p["email"]))

    def remove_selected_participant(self):
        # Remove um participante selecionado da treeview e da lista interna
        sel = self.part_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um participante para remover.")
            return
        idx = self.part_tree.index(sel[0])
        self.part_tree.delete(sel[0])
        del self.participants[idx]

    def add_item_dialog(self):
        # Abre o diálogo para adicionar um item ao novo projeto (ainda em criação)
        dlg = AddItemDialog(self)
        self.wait_window(dlg) # Espera o diálogo AddItemDialog fechar
        if dlg.result:
            item = dlg.result
            if any(i['desc'].lower() == item['desc'].lower() for i in self.items):
                messagebox.showwarning("Duplicado", "Já existe um item com esta descrição neste projeto. Considere editar a quantidade ou usar uma descrição diferente.")
                return
            self.items.append(item)
            self.items_tree.insert("", "end", values=(item['desc'], item['qty'], item['unit'], item['unit_price'], item['total']))

    def remove_selected_item(self):
        # Remove um item selecionado da treeview e da lista interna
        sel = self.items_tree.selection()
        if not sel:
            messagebox.showwarning("Aviso", "Selecione um item para remover.")
            return
        idx = self.items_tree.index(sel[0])
        self.items_tree.delete(sel[0])
        del self.items[idx]

    def save_project(self):
        # Valida todos os campos do novo projeto antes de retornar os dados
        name = self.entry_name.get().strip() # Não substitui espaços aqui, será feito no save_project_spreadsheet para nome do arquivo
        
        if not name:
            messagebox.showerror("Erro", "O Nome do Projeto é obrigatório.")
            return
        
        # Normaliza o nome do arquivo, mas mantém o nome original para exibição
        project_name_for_file = name.replace(" ", "_")

        manager = self.entry_manager.get().strip()
        email = self.entry_manager_email.get().strip()
        start_date = self.entry_start.get().strip()
        end_date = self.entry_end.get().strip()
        
        try:
            est_cost = float(self.entry_cost.get().strip())
        except ValueError:
            messagebox.showerror("Erro", "O Custo Estimado deve ser um número válido (ex: 100.00).")
            return
        
        if not manager or not email or not start_date or not end_date:
            messagebox.showerror("Erro", "Todos os campos de informação do projeto (Gerente, Email, Datas) são obrigatórios.")
            return
        
        if not self.items:
            messagebox.showerror("Erro", "Adicione pelo menos um item ao projeto.")
            return

        # Verifica se o nome do projeto (para arquivo) já existe
        existing_project_files = [f.lower() for f in os.listdir() if f.startswith("project_") and f.endswith(".xlsx") or f.startswith("project_")] # Inclui verificação sem .xlsx para pastas
        if f"project_{project_name_for_file.lower()}.xlsx" in existing_project_files:
            messagebox.showwarning("Nome de Projeto Duplicado", f"Já existe um arquivo para um projeto com o nome '{name}'. Por favor, escolha um nome diferente ou exclua o projeto existente.")
            return

        # Coleta todas as informações em um dicionário
        info = {
            "project_name": name, # Mantém o nome original com espaços para exibição
            "manager": manager,
            "manager_email": email,
            "start_date": start_date,
            "end_date": end_date,
            "est_cost": est_cost,
            "participants": self.participants
        }
        
        # Armazena os resultados e fecha o diálogo
        self.result = {"info": info, "items": self.items}
        self.destroy() # Fecha o diálogo


class AddParticipantDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Adicionar Participante")
        self.result = None
        
        # Ajuste o tamanho inicial para ser menor e centralizado
        self.initial_width = 300
        self.initial_height = 160
        self.center_window(self.initial_width, self.initial_height, parent)
        self.transient(parent)
        self.grab_set()

        frm = ttk.Frame(self)
        frm.pack(padx=10, pady=10, fill="both", expand=True)

        ttk.Label(frm, text="Nome:").pack(anchor="w")
        self.entry_name = ttk.Entry(frm)
        self.entry_name.pack(fill="x", pady=2)

        ttk.Label(frm, text="Email:").pack(anchor="w", pady=(10,0))
        self.entry_email = ttk.Entry(frm)
        self.entry_email.pack(fill="x", pady=2)

        btn_frame = ttk.Frame(frm)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Adicionar", command=self.on_add).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=self.destroy).pack(side="left", padx=5)
    
    def center_window(self, width, height, parent):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width / 2) - (width / 2)
        y = (screen_height / 2) - (height / 2)
        self.geometry(f'{int(width)}x{int(height)}+{int(x)}+{int(y)}')

    def on_add(self):
        name = self.entry_name.get().strip()
        email = self.entry_email.get().strip()
        if not name or not email:
            messagebox.showerror("Erro", "Nome e e-mail são obrigatórios.")
            return
        if "@" not in email or "." not in email.split("@")[-1] or len(email.split("@")[0]) == 0:
            messagebox.showerror("Erro", "Formato de e-mail inválido.")
            return
        self.result = {"name": name, "email": email}
        self.destroy()

class AddItemDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Adicionar Item")
        self.result = None
        
        # Ajuste o tamanho inicial para ser menor e centralizado
        self.initial_width = 350
        self.initial_height = 280
        self.center_window(self.initial_width, self.initial_height, parent)
        self.transient(parent)
        self.grab_set()

        frm = ttk.Frame(self)
        frm.pack(padx=10, pady=10, fill="both", expand=True)

        ttk.Label(frm, text="Descrição:").pack(anchor="w")
        self.entry_desc = ttk.Entry(frm)
        self.entry_desc.pack(fill="x", pady=2)

        ttk.Label(frm, text="Quantidade:").pack(anchor="w", pady=(10,0))
        self.entry_qty = ttk.Entry(frm)
        self.entry_qty.pack(fill="x", pady=2)

        ttk.Label(frm, text="Unidade:").pack(anchor="w", pady=(10,0))
        self.entry_unit = ttk.Entry(frm)
        self.entry_unit.pack(fill="x", pady=2)

        ttk.Label(frm, text="Preço Unitário (R$):").pack(anchor="w", pady=(10,0))
        self.entry_unit_price = ttk.Entry(frm)
        self.entry_unit_price.pack(fill="x", pady=2)

        btn_frame = ttk.Frame(frm)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Adicionar", command=self.on_add).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=self.destroy).pack(side="left", padx=5)
    
    def center_window(self, width, height, parent):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width / 2) - (width / 2)
        y = (screen_height / 2) - (height / 2)
        self.geometry(f'{int(width)}x{int(height)}+{int(x)}+{int(y)}')

    def on_add(self):
        desc = self.entry_desc.get().strip()
        if not desc:
            messagebox.showerror("Erro", "Descrição é obrigatória.")
            return
        try:
            qty = float(self.entry_qty.get().strip())
            if qty <= 0:
                raise ValueError()
        except ValueError:
            messagebox.showerror("Erro", "Quantidade deve ser um número positivo.")
            return
        unit = self.entry_unit.get().strip()
        if not unit:
            messagebox.showerror("Erro", "Unidade é obrigatória.")
            return
        try:
            unit_price = float(self.entry_unit_price.get().strip())
            if unit_price < 0:
                raise ValueError()
        except ValueError:
            messagebox.showerror("Erro", "Preço unitário deve ser um número não-negativo.")
            return
        total = qty * unit_price
        self.result = {
            "desc": desc,
            "qty": qty,
            "unit": unit,
            "unit_price": unit_price,
            "total": total
        }
        self.destroy()

if __name__ == "__main__":
    app = ProjectManagerApp()
    app.mainloop()
